"""File loaders for xlsx, csv, and pdf input."""

import os
import pandas as pd
from llm_detector.compat import HAS_PYPDF

if HAS_PYPDF:
    from pypdf import PdfReader


def _col_letter_to_index(spec):
    """Convert a column letter (A-Z) or 1-based number string to a 0-based index.

    Returns the integer index if *spec* is a positional reference, else None.
    Supports single letters A-Z (case-insensitive) and positive integer strings.
    """
    s = str(spec).strip()
    if len(s) == 1 and s.upper() in 'ABCDEFGHIJKLMNOPQRSTUVWXYZ':
        return ord(s.upper()) - ord('A')
    if s.isdigit() and int(s) >= 1:
        return int(s) - 1  # convert 1-based to 0-based
    return None


def load_xlsx(filepath, sheet=None, prompt_col='prompt', id_col='task_id',
              occ_col='occupation', attempter_col='attempter_name',
              stage_col='pipeline_stage_name',
              attempter_email_col='', reviewer_col='', reviewer_email_col=''):
    """Load tasks from an xlsx file. Returns list of dicts.

    Column parameters accept either a column header name (matched case-insensitively
    with fuzzy fallback) or a positional reference: a single letter A–Z (e.g. ``'A'``
    for the first column) or a 1-based integer string (e.g. ``'1'`` for the first
    column).  Positional references take priority over name matching.
    """
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True)

    if sheet:
        ws = wb[sheet]
    else:
        for name in ['FullTaskX', 'Full Task Connected', 'Claim Sheet', 'Sample List']:
            if name in wb.sheetnames:
                ws = wb[name]
                break
        else:
            ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()

    if not rows:
        return []

    headers = [str(h).strip().lower() if h else '' for h in rows[0]]
    n_cols = len(headers)

    def find_col(candidates):
        # Positional spec takes priority (first matching candidate wins)
        for c in candidates:
            idx = _col_letter_to_index(c)
            if idx is not None and idx < n_cols:
                return idx
        # Exact name match
        for c in candidates:
            cl = c.lower()
            for i, h in enumerate(headers):
                if cl == h:
                    return i
        # Substring match (skip bare 'id' to avoid false positives)
        for c in candidates:
            cl = c.lower()
            if cl == 'id':
                continue
            for i, h in enumerate(headers):
                if cl in h:
                    return i
        return None

    prompt_idx = find_col([prompt_col, 'prompt', 'text', 'content'])
    id_idx = find_col([id_col, 'task_id', 'id'])
    occ_idx = find_col([occ_col, 'occupation', 'occ'])
    att_idx = find_col([attempter_col, 'attempter', 'claimed_by', 'claimed by',
                        'fellow name', 'fellow_name', 'author', 'name'])
    stage_idx = find_col([stage_col, 'stage', 'pipeline_stage'])
    att_email_idx = find_col([attempter_email_col]) if attempter_email_col else None
    rev_idx = find_col([reviewer_col]) if reviewer_col else None
    rev_email_idx = find_col([reviewer_email_col]) if reviewer_email_col else None

    if prompt_idx is None:
        print(f"ERROR: Could not find prompt column. Headers: {headers}")
        return []

    tasks = []
    for row in rows[1:]:
        if not row or len(row) <= prompt_idx:
            continue
        prompt = str(row[prompt_idx]).strip() if row[prompt_idx] else ''
        if len(prompt) < 50:
            continue

        task = {
            'prompt': prompt,
            'task_id': str(row[id_idx])[:20] if id_idx is not None and row[id_idx] else '',
            'occupation': str(row[occ_idx]) if occ_idx is not None and row[occ_idx] else '',
            'attempter': str(row[att_idx]) if att_idx is not None and row[att_idx] else '',
            'stage': str(row[stage_idx]) if stage_idx is not None and row[stage_idx] else '',
        }
        if att_email_idx is not None and att_email_idx < len(row):
            task['attempter_email'] = str(row[att_email_idx]) if row[att_email_idx] else ''
        if rev_idx is not None and rev_idx < len(row):
            task['reviewer'] = str(row[rev_idx]) if row[rev_idx] else ''
        if rev_email_idx is not None and rev_email_idx < len(row):
            task['reviewer_email'] = str(row[rev_email_idx]) if row[rev_email_idx] else ''
        tasks.append(task)

    return tasks


def load_csv(filepath, prompt_col='prompt', id_col='task_id',
             occ_col='occupation', attempter_col='attempter_name',
             stage_col='pipeline_stage_name',
             attempter_email_col='', reviewer_col='', reviewer_email_col=''):
    """Load tasks from CSV.

    Column parameters accept either a column header name (matched
    case-insensitively with fuzzy fallback) or a positional reference: a single
    letter A–Z (e.g. ``'A'``) or a 1-based integer string (e.g. ``'1'``).
    """
    df = pd.read_csv(filepath)
    df = df.fillna('')

    col_list = list(df.columns)
    col_map = {c.lower().strip(): c for c in col_list}

    def resolve_col(*candidates):
        # Positional spec takes priority
        for c in candidates:
            idx = _col_letter_to_index(c)
            if idx is not None and idx < len(col_list):
                return col_list[idx]
        # Exact name match
        for c in candidates:
            key = c.lower().strip()
            if key in col_map:
                return col_map[key]
        # Substring match
        for c in candidates:
            key = c.lower().strip()
            if key == 'id':
                continue
            for mapped_key, actual in col_map.items():
                if key in mapped_key:
                    return actual
        return None

    prompt_actual = resolve_col(prompt_col, 'prompt', 'text', 'content')
    id_actual = resolve_col(id_col, 'task_id', 'id')
    occ_actual = resolve_col(occ_col, 'occupation', 'occ')
    att_actual = resolve_col(attempter_col, 'attempter_name', 'attempter', 'claimed_by',
                             'fellow name', 'fellow_name', 'author', 'name')
    stage_actual = resolve_col(stage_col, 'pipeline_stage_name', 'stage')
    att_email_actual = resolve_col(attempter_email_col) if attempter_email_col else None
    rev_actual = resolve_col(reviewer_col) if reviewer_col else None
    rev_email_actual = resolve_col(reviewer_email_col) if reviewer_email_col else None

    if prompt_actual is None:
        print(f"ERROR: Could not find prompt column. Columns: {list(df.columns)}")
        return []

    tasks = []
    for _, row in df.iterrows():
        prompt = str(row.get(prompt_actual, ''))
        if len(prompt) < 50:
            continue
        task = {
            'prompt': prompt,
            'task_id': str(row.get(id_actual, ''))[:20] if id_actual else '',
            'occupation': str(row.get(occ_actual, '')) if occ_actual else '',
            'attempter': str(row.get(att_actual, '')) if att_actual else '',
            'stage': str(row.get(stage_actual, '')) if stage_actual else '',
        }
        if att_email_actual:
            task['attempter_email'] = str(row.get(att_email_actual, ''))
        if rev_actual:
            task['reviewer'] = str(row.get(rev_actual, ''))
        if rev_email_actual:
            task['reviewer_email'] = str(row.get(rev_email_actual, ''))
        tasks.append(task)
    return tasks


def load_pdf(filepath):
    """Load text from PDF file. Each page becomes a separate task."""
    if not HAS_PYPDF:
        print("ERROR: pypdf not installed. Run: pip install pypdf")
        return []

    reader = PdfReader(filepath)
    tasks = []
    for i, page in enumerate(reader.pages):
        text = page.extract_text()
        if text and len(text.strip()) >= 50:
            tasks.append({
                'prompt': text.strip(),
                'task_id': f"page_{i+1}",
                'occupation': '',
                'attempter': '',
                'stage': '',
            })

    if not tasks:
        full_text = '\n'.join(
            page.extract_text() for page in reader.pages
            if page.extract_text()
        ).strip()
        if len(full_text) >= 50:
            tasks.append({
                'prompt': full_text,
                'task_id': 'full_document',
                'occupation': '',
                'attempter': '',
                'stage': '',
            })

    return tasks
