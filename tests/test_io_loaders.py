"""Tests for io.py file loaders — target 85% coverage."""

import sys
import os
import csv
import tempfile
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from llm_detector.io import load_xlsx, load_csv, load_pdf, _col_letter_to_index
from llm_detector.compat import HAS_PYPDF

PASSED = 0
FAILED = 0

LONG_PROMPT = "This is a test prompt that is definitely longer than fifty characters."


def check(label, condition, detail=""):
    global PASSED, FAILED
    if condition:
        PASSED += 1
        print(f"  [PASS] {label}")
    else:
        FAILED += 1
        print(f"  [FAIL] {label}  -- {detail}")


# ── _col_letter_to_index ──────────────────────────────────────────────────────

def test_col_letter_to_index_letters():
    """Letter A-Z are converted to 0-based indices."""
    print("\n-- COL LETTER TO INDEX (letters) --")
    check("A -> 0", _col_letter_to_index('A') == 0)
    check("B -> 1", _col_letter_to_index('B') == 1)
    check("Z -> 25", _col_letter_to_index('Z') == 25)
    check("a -> 0 (lowercase)", _col_letter_to_index('a') == 0)
    check("z -> 25 (lowercase)", _col_letter_to_index('z') == 25)


def test_col_letter_to_index_numbers():
    """Positive integer strings are converted to 0-based indices."""
    print("\n-- COL LETTER TO INDEX (numbers) --")
    check("'1' -> 0", _col_letter_to_index('1') == 0)
    check("'2' -> 1", _col_letter_to_index('2') == 1)
    check("'10' -> 9", _col_letter_to_index('10') == 9)


def test_col_letter_to_index_invalid():
    """Invalid specs return None."""
    print("\n-- COL LETTER TO INDEX (invalid) --")
    check("'invalid' -> None", _col_letter_to_index('invalid') is None)
    check("'0' -> None (0 is not 1-based)", _col_letter_to_index('0') is None)
    check("'AB' -> None (multi-letter not supported)",
          _col_letter_to_index('AB') is None)
    check("'' -> None", _col_letter_to_index('') is None)


# ── load_xlsx ─────────────────────────────────────────────────────────────────

def _make_xlsx(headers, rows, sheet_name=None):
    """Helper: create a temp xlsx file. Returns filepath."""
    wb = openpyxl.Workbook()
    ws = wb.active if sheet_name is None else wb.create_sheet(sheet_name)
    if sheet_name is not None:
        # Remove the default 'Sheet' to avoid confusion
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
    ws.append(headers)
    for row in rows:
        ws.append(row)
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        filepath = f.name
    wb.save(filepath)
    wb.close()
    return filepath


def test_load_xlsx_basic():
    """load_xlsx loads a simple file with standard column names."""
    print("\n-- LOAD XLSX BASIC --")
    filepath = _make_xlsx(
        ['prompt', 'task_id', 'occupation'],
        [[LONG_PROMPT, 'task1', 'teacher']],
    )
    try:
        tasks = load_xlsx(filepath)
        check("1 task loaded", len(tasks) == 1, f"got {len(tasks)}")
        check("prompt correct", LONG_PROMPT in tasks[0]['prompt'])
        check("task_id correct", tasks[0]['task_id'] == 'task1')
        check("occupation correct", tasks[0]['occupation'] == 'teacher')
    finally:
        os.unlink(filepath)


def test_load_xlsx_default_sheet_fulltaskx():
    """load_xlsx selects FullTaskX sheet automatically."""
    print("\n-- LOAD XLSX FULLTASKX SHEET --")
    wb = openpyxl.Workbook()
    # Create a dummy default sheet and the preferred sheet
    ws_other = wb.active
    ws_other.title = 'OtherSheet'
    ws_other.append(['ignored_col'])

    ws_pref = wb.create_sheet('FullTaskX')
    ws_pref.append(['prompt', 'task_id'])
    ws_pref.append([LONG_PROMPT, 'taskX'])

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        filepath = f.name
    wb.save(filepath)
    wb.close()
    try:
        tasks = load_xlsx(filepath)
        check("FullTaskX sheet auto-selected", len(tasks) == 1, f"got {len(tasks)}")
        check("Task from FullTaskX", tasks[0].get('task_id') == 'taskX')
    finally:
        os.unlink(filepath)


def test_load_xlsx_fallback_first_sheet():
    """load_xlsx falls back to first sheet when no known sheet name found."""
    print("\n-- LOAD XLSX FALLBACK FIRST SHEET --")
    filepath = _make_xlsx(
        ['prompt', 'task_id'],
        [[LONG_PROMPT, 'fallback_task']],
        sheet_name='CustomSheet',
    )
    try:
        tasks = load_xlsx(filepath)
        check("Fallback to first sheet", len(tasks) == 1, f"got {len(tasks)}")
        check("Task from first sheet",
              tasks[0].get('task_id') == 'fallback_task')
    finally:
        os.unlink(filepath)


def test_load_xlsx_positional_columns():
    """load_xlsx accepts positional column references (A, B, C)."""
    print("\n-- LOAD XLSX POSITIONAL COLUMNS --")
    filepath = _make_xlsx(
        ['text_content', 'unique_id', 'job_type'],
        [[LONG_PROMPT, 'pos_task', 'engineer']],
    )
    try:
        tasks = load_xlsx(filepath, prompt_col='A', id_col='B', occ_col='C')
        check("Positional: 1 task loaded", len(tasks) == 1, f"got {len(tasks)}")
        check("Positional A (prompt)", LONG_PROMPT in tasks[0]['prompt'])
        check("Positional B (id)", tasks[0]['task_id'] == 'pos_task')
        check("Positional C (occ)", tasks[0]['occupation'] == 'engineer')
    finally:
        os.unlink(filepath)


def test_load_xlsx_positional_numeric():
    """load_xlsx accepts 1-based numeric positional references."""
    print("\n-- LOAD XLSX POSITIONAL NUMERIC --")
    filepath = _make_xlsx(
        ['text_content', 'unique_id', 'job_type'],
        [[LONG_PROMPT, 'num_task', 'doctor']],
    )
    try:
        tasks = load_xlsx(filepath, prompt_col='1', id_col='2', occ_col='3')
        check("Numeric positional: 1 task loaded", len(tasks) == 1)
        check("Numeric positional A (prompt)", LONG_PROMPT in tasks[0]['prompt'])
    finally:
        os.unlink(filepath)


def test_load_xlsx_fuzzy_column_matching():
    """load_xlsx finds columns via substring matching."""
    print("\n-- LOAD XLSX FUZZY MATCHING --")
    filepath = _make_xlsx(
        ['user_submitted_prompt', 'unique_task_identifier', 'employee_occupation'],
        [[LONG_PROMPT, 'fuzz_task', 'nurse']],
    )
    try:
        tasks = load_xlsx(filepath)
        check("Fuzzy match: 1 task loaded", len(tasks) == 1, f"got {len(tasks)}")
        check("Fuzzy prompt matched", LONG_PROMPT in tasks[0]['prompt'])
    finally:
        os.unlink(filepath)


def test_load_xlsx_short_prompt_filtered():
    """load_xlsx filters out prompts shorter than 50 characters."""
    print("\n-- LOAD XLSX SHORT PROMPT FILTER --")
    filepath = _make_xlsx(
        ['prompt'],
        [['Too short.'], [LONG_PROMPT]],
    )
    try:
        tasks = load_xlsx(filepath)
        check("Short prompt filtered", len(tasks) == 1, f"got {len(tasks)}")
        check("Long prompt kept", LONG_PROMPT in tasks[0]['prompt'])
    finally:
        os.unlink(filepath)


def test_load_xlsx_no_prompt_column():
    """load_xlsx returns empty list when prompt column is not found."""
    print("\n-- LOAD XLSX NO PROMPT COLUMN --")
    filepath = _make_xlsx(
        ['col1', 'col2'],
        [['value1', 'value2']],
    )
    try:
        tasks = load_xlsx(filepath, prompt_col='nonexistent_col_xyz')
        check("No prompt col -> empty list", len(tasks) == 0, f"got {len(tasks)}")
    finally:
        os.unlink(filepath)


def test_load_xlsx_empty_workbook():
    """load_xlsx handles workbook with no data rows."""
    print("\n-- LOAD XLSX EMPTY --")
    filepath = _make_xlsx(['prompt', 'task_id'], [])
    try:
        tasks = load_xlsx(filepath)
        check("Empty workbook -> empty list", len(tasks) == 0, f"got {len(tasks)}")
    finally:
        os.unlink(filepath)


def test_load_xlsx_explicit_sheet():
    """load_xlsx uses explicit sheet name when provided."""
    print("\n-- LOAD XLSX EXPLICIT SHEET --")
    wb = openpyxl.Workbook()
    ws_main = wb.active
    ws_main.title = 'Main'
    ws_main.append(['prompt'])
    ws_main.append(['Wrong sheet prompt that is also over fifty characters long.'])

    ws_alt = wb.create_sheet('AltSheet')
    ws_alt.append(['prompt', 'task_id'])
    ws_alt.append([LONG_PROMPT, 'alt_task'])

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        filepath = f.name
    wb.save(filepath)
    wb.close()
    try:
        tasks = load_xlsx(filepath, sheet='AltSheet')
        check("Explicit sheet selected", len(tasks) == 1, f"got {len(tasks)}")
        check("Task from AltSheet", tasks[0].get('task_id') == 'alt_task')
    finally:
        os.unlink(filepath)


def test_load_xlsx_attempter_and_stage():
    """load_xlsx loads attempter and stage columns."""
    print("\n-- LOAD XLSX ATTEMPTER & STAGE --")
    filepath = _make_xlsx(
        ['prompt', 'task_id', 'occupation', 'attempter_name', 'pipeline_stage_name'],
        [[LONG_PROMPT, 'task1', 'dev', 'alice', 'review']],
    )
    try:
        tasks = load_xlsx(filepath)
        check("attempter loaded", tasks[0].get('attempter') == 'alice')
        check("stage loaded", tasks[0].get('stage') == 'review')
    finally:
        os.unlink(filepath)


# ── load_csv ──────────────────────────────────────────────────────────────────

def _make_csv(headers, rows):
    """Helper: create a temp CSV file. Returns filepath."""
    with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False,
                                     newline='') as f:
        filepath = f.name
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)
    return filepath


def test_load_csv_basic():
    """load_csv loads a simple CSV with standard column names."""
    print("\n-- LOAD CSV BASIC --")
    filepath = _make_csv(
        ['prompt', 'task_id', 'occupation'],
        [[LONG_PROMPT, 'csv_task', 'teacher']],
    )
    try:
        tasks = load_csv(filepath)
        check("CSV: 1 task loaded", len(tasks) == 1, f"got {len(tasks)}")
        check("CSV prompt correct", LONG_PROMPT in tasks[0]['prompt'])
        check("CSV task_id correct", tasks[0]['task_id'] == 'csv_task')
        check("CSV occupation correct", tasks[0]['occupation'] == 'teacher')
    finally:
        os.unlink(filepath)


def test_load_csv_positional_letter():
    """load_csv accepts positional column references (A, B, C)."""
    print("\n-- LOAD CSV POSITIONAL LETTER --")
    filepath = _make_csv(
        ['text_content', 'unique_id', 'job_type'],
        [[LONG_PROMPT, 'pos_csv', 'engineer']],
    )
    try:
        tasks = load_csv(filepath, prompt_col='A', id_col='B', occ_col='C')
        check("CSV positional A: 1 task", len(tasks) == 1)
        check("CSV positional A prompt", LONG_PROMPT in tasks[0]['prompt'])
        check("CSV positional B id", tasks[0]['task_id'] == 'pos_csv')
    finally:
        os.unlink(filepath)


def test_load_csv_positional_numeric():
    """load_csv accepts numeric positional references."""
    print("\n-- LOAD CSV POSITIONAL NUMERIC --")
    filepath = _make_csv(
        ['text_content', 'unique_id'],
        [[LONG_PROMPT, 'num_csv']],
    )
    try:
        tasks = load_csv(filepath, prompt_col='1', id_col='2')
        check("CSV numeric positional: 1 task", len(tasks) == 1)
        check("Numeric position prompt loaded", LONG_PROMPT in tasks[0]['prompt'])
    finally:
        os.unlink(filepath)


def test_load_csv_fuzzy_matching():
    """load_csv matches columns by substring."""
    print("\n-- LOAD CSV FUZZY MATCHING --")
    filepath = _make_csv(
        ['user_submitted_prompt', 'task_identifier'],
        [[LONG_PROMPT, 'fuzz_csv']],
    )
    try:
        tasks = load_csv(filepath)
        check("CSV fuzzy match: 1 task", len(tasks) == 1, f"got {len(tasks)}")
        check("CSV fuzzy prompt matched", LONG_PROMPT in tasks[0]['prompt'])
    finally:
        os.unlink(filepath)


def test_load_csv_missing_prompt_column():
    """load_csv returns empty list when prompt column is not found."""
    print("\n-- LOAD CSV MISSING PROMPT --")
    filepath = _make_csv(
        ['col1', 'col2'],
        [['value1', 'value2']],
    )
    try:
        tasks = load_csv(filepath, prompt_col='nonexistent_xyz')
        check("Missing prompt col -> empty list", len(tasks) == 0, f"got {len(tasks)}")
    finally:
        os.unlink(filepath)


def test_load_csv_short_prompt_filtered():
    """load_csv filters out prompts shorter than 50 characters."""
    print("\n-- LOAD CSV SHORT PROMPT FILTER --")
    filepath = _make_csv(
        ['prompt'],
        [['Short.'], [LONG_PROMPT]],
    )
    try:
        tasks = load_csv(filepath)
        check("Short CSV prompt filtered", len(tasks) == 1, f"got {len(tasks)}")
        check("Long CSV prompt kept", LONG_PROMPT in tasks[0]['prompt'])
    finally:
        os.unlink(filepath)


def test_load_csv_attempter_and_stage():
    """load_csv loads attempter and stage columns."""
    print("\n-- LOAD CSV ATTEMPTER & STAGE --")
    filepath = _make_csv(
        ['prompt', 'task_id', 'occupation', 'attempter_name', 'pipeline_stage_name'],
        [[LONG_PROMPT, 'task1', 'dev', 'bob', 'draft']],
    )
    try:
        tasks = load_csv(filepath)
        check("CSV attempter loaded", tasks[0].get('attempter') == 'bob')
        check("CSV stage loaded", tasks[0].get('stage') == 'draft')
    finally:
        os.unlink(filepath)


# ── load_pdf ──────────────────────────────────────────────────────────────────

def test_load_pdf_no_pypdf():
    """load_pdf returns empty list when pypdf is not installed."""
    print("\n-- LOAD PDF NO PYPDF --")
    if not HAS_PYPDF:
        tasks = load_pdf('dummy_nonexistent.pdf')
        check("load_pdf without pypdf -> empty list",
              len(tasks) == 0, f"got {len(tasks)}")
    else:
        # pypdf is installed; just verify the function exists and handles bad path
        try:
            load_pdf('nonexistent_file_xyzzy.pdf')
            check("PDF with bad path raises exception", False)
        except Exception:
            check("PDF with bad path raises exception (expected)", True)


if __name__ == '__main__':
    print("=" * 70)
    print("IO Loaders Tests")
    print("=" * 70)

    test_col_letter_to_index_letters()
    test_col_letter_to_index_numbers()
    test_col_letter_to_index_invalid()
    test_load_xlsx_basic()
    test_load_xlsx_default_sheet_fulltaskx()
    test_load_xlsx_fallback_first_sheet()
    test_load_xlsx_positional_columns()
    test_load_xlsx_positional_numeric()
    test_load_xlsx_fuzzy_column_matching()
    test_load_xlsx_short_prompt_filtered()
    test_load_xlsx_no_prompt_column()
    test_load_xlsx_empty_workbook()
    test_load_xlsx_explicit_sheet()
    test_load_xlsx_attempter_and_stage()
    test_load_csv_basic()
    test_load_csv_positional_letter()
    test_load_csv_positional_numeric()
    test_load_csv_fuzzy_matching()
    test_load_csv_missing_prompt_column()
    test_load_csv_short_prompt_filtered()
    test_load_csv_attempter_and_stage()
    test_load_pdf_no_pypdf()

    print(f"\n{'=' * 70}")
    print(f"RESULTS: {PASSED} passed, {FAILED} failed")
    print(f"{'=' * 70}")
    if FAILED > 0:

        sys.exit(1)
